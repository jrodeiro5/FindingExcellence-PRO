"""
Excel file processing module.

This module handles the reading and analysis of Excel files.
"""

import os
import logging
import openpyxl
import pandas as pd
import importlib

class ExcelProcessor:
    """
    Processes Excel files for content searching.
    """
    
    SUPPORTED_EXTENSIONS = ('.xls', '.xlsx', '.xlsm')
    
    @staticmethod
    def diagnose_excel_file(file_path):
        """
        Diagnose and attempt to open an Excel file with various strategies.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            tuple: (excel_data, error_message)
        """
        try:
            # 1. Verify file exists and has size
            if not os.path.exists(file_path):
                return None, "File does not exist"
            
            file_size = os.path.getsize(file_path)
            if file_size == 0:
                return None, "File is empty (0 bytes)"
                
            # 2. Check file extension
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in ExcelProcessor.SUPPORTED_EXTENSIONS:
                return None, f"Unexpected file extension: {ext}"
            
            # Log diagnostic info    
            logging.info(f"Diagnosing Excel file: {file_path} with extension {ext} and size {file_size} bytes")
                
            # 3. Try different strategies based on file type
            if ext == '.xls':
                # Use pandas with xlrd for .xls files
                try:
                    # Import xlrd here to avoid unnecessary dependencies
                    if importlib.util.find_spec("xlrd") is None:
                        return None, "xlrd module not available for reading .xls files"
                        
                    # Try to read with xlrd engine
                    logging.info(f"Attempting to read .xls file with xlrd engine: {file_path}")
                    df = pd.read_excel(file_path, engine='xlrd', sheet_name=None)
                    return df, None
                except Exception as e:
                    error_msg = f"Failed to read .xls file: {str(e)}"
                    logging.error(error_msg)
                    return None, error_msg
            else:  # .xlsx or .xlsm
                # Try different strategies for xlsx/xlsm files
                errors = []
                
                # Strategy 1: pandas with openpyxl engine
                try:
                    logging.info(f"Attempting to read file with pandas+openpyxl: {file_path}")
                    df = pd.read_excel(file_path, engine='openpyxl', sheet_name=None)
                    return df, None
                except Exception as e:
                    error_msg = f"Pandas+openpyxl error: {str(e)}"
                    logging.error(error_msg)
                    errors.append(error_msg)
                
                # Strategy 2: direct openpyxl
                try:
                    logging.info(f"Attempting to read file with direct openpyxl: {file_path}")
                    # Adding strict=False to help with problematic files
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True, strict=False)
                    return wb, None
                except Exception as e:
                    error_msg = f"Direct openpyxl error: {str(e)}"
                    logging.error(error_msg)
                    errors.append(error_msg)
                
                # Strategy 3: Fallback for corrupted Excel files
                try:
                    # More primitive approach for problematic files
                    # Read as text and check for keywords
                    logging.info(f"Attempting fallback text-based search for: {file_path}")
                    with open(file_path, 'rb') as f:
                        raw_content = f.read()
                        # Convert to string, ignore errors
                        text_content = raw_content.decode('utf-8', errors='ignore')
                    
                    # Create wrapper to mimic Excel structure
                    class SimpleTextWrapper:
                        def __init__(self, content):
                            self.content = content
                            self.sheetnames = ['TextContent']
                        
                        def __getitem__(self, sheet_name):
                            # Simple wrapper that generates mock cells
                            class MockSheet:
                                def __init__(self, content):
                                    self.content = content
                                
                                def iter_rows(self):
                                    # Return a single row with a single cell
                                    class MockCell:
                                        def __init__(self, value):
                                            self.value = value
                                    
                                    return [[MockCell(self.content)]]
                            
                            return MockSheet(self.content)
                    
                    # If we have text content, use it as a last resort
                    if text_content and len(text_content) > 0:
                        logging.info(f"Using text-based fallback search for: {file_path}")
                        return SimpleTextWrapper(text_content), None
                except Exception as e:
                    error_msg = f"Fallback text approach error: {str(e)}"
                    logging.error(error_msg)
                    errors.append(error_msg)
                
                # If nothing worked, return detailed error
                return None, "\n".join(["Multiple read attempts failed:"] + errors)
                
        except Exception as general_error:
            return None, f"Diagnosis error: {str(general_error)}"
    
    @staticmethod
    def search_content(file_path, keywords, case_sensitive=False, cancel_event=None):
        """
        Search for keywords in an Excel file's content.
        
        Args:
            file_path: Path to the Excel file
            keywords: List of keywords to search for
            case_sensitive: Whether to perform case-sensitive search
            cancel_event: Optional threading event for cancellation
            
        Returns:
            list: List of matches found
        """
        file_results = []
        
        try:
            # Log file details
            try:
                file_size = os.path.getsize(file_path)
                logging.info(f"Processing file: {file_path}, Size: {file_size} bytes")
            except Exception as e:
                logging.warning(f"Unable to get file size for {file_path}: {e}")
            
            # Prepare keywords based on case sensitivity
            processed_keywords = keywords if case_sensitive else [k.lower() for k in keywords]

            def check_cell(cell_val_str):
                val_to_check = cell_val_str if case_sensitive else cell_val_str.lower()
                for idx, kw in enumerate(processed_keywords):
                    if kw in val_to_check:
                        return keywords[idx]  # Return original keyword
                return None

            # Diagnose and open the file
            excel_data, error_msg = ExcelProcessor.diagnose_excel_file(file_path)
            
            if error_msg:
                # If diagnosis failed, log and return error
                logging.error(f"Excel diagnosis error for {file_path}: {error_msg}")
                file_results.append({'error': f"Error processing file: {error_msg}", 'file_path': file_path})
                return file_results
                
            # Process based on the type of data returned
            if isinstance(excel_data, dict):  # pandas DataFrame dict
                # Process pandas DataFrames
                for sheet_name, df in excel_data.items():
                    # Check for cancellation before processing each sheet
                    if cancel_event and cancel_event.is_set():
                        logging.debug(f"Content search cancelled while processing sheet {sheet_name} in {file_path}")
                        break
                        
                    # Iterate over all cells
                    for row_idx, row_series in df.iterrows():
                        # Check for cancellation every few rows for responsiveness
                        if cancel_event and cancel_event.is_set() and row_idx % 50 == 0:
                            logging.debug(f"Content search cancelled at row {row_idx} in {file_path}")
                            break
                            
                        for col_idx, cell_value in enumerate(row_series, start=1):
                            # Process only if not NaN
                            if pd.notna(cell_value):
                                cell_value_str = str(cell_value)
                                found_keyword = check_cell(cell_value_str)
                                if found_keyword:
                                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                                    # In pandas rows start at 0, add 1 for A1 notation
                                    file_results.append({
                                        'keyword': found_keyword, 
                                        'sheet': sheet_name,
                                        'cell': f"{col_letter}{row_idx+1}",
                                        'value': cell_value_str
                                    })
                                    
            elif hasattr(excel_data, 'sheetnames'):  # openpyxl Workbook
                # Process openpyxl workbook
                for sheet_name in excel_data.sheetnames:
                    # Check for cancellation before processing each sheet
                    if cancel_event and cancel_event.is_set():
                        logging.debug(f"Content search cancelled while processing sheet {sheet_name} in {file_path}")
                        break
                        
                    sheet = excel_data[sheet_name]
                    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                        # Check for cancellation every few rows for responsiveness
                        if cancel_event and cancel_event.is_set() and row_idx % 50 == 0:
                            logging.debug(f"Content search cancelled at row {row_idx} in {file_path}")
                            break
                            
                        for col_idx, cell in enumerate(row, start=1):
                            cell_value_str = str(cell.value) if cell.value is not None else ""
                            found_keyword = check_cell(cell_value_str)
                            if found_keyword:
                                file_results.append({
                                    'keyword': found_keyword, 
                                    'sheet': sheet_name,
                                    'cell': f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}",
                                    'value': cell_value_str
                                })
            else:
                # Unexpected object type
                raise Exception(f"Unexpected data type returned from file diagnosis: {type(excel_data)}")
                    
        except Exception as e:
            logging.error(f"Error processing content of {file_path}: {e}", exc_info=True)
            # Add error marker for UI
            file_results.append({
                'error': f"Error processing file: {str(e)}", 
                'file_path': file_path
            })
        
        return file_results
