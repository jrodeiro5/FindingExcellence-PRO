"""Results panel using native Tkinter Treeview for professional table display."""

import csv
import os
import subprocess
import sys
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, ttk
from typing import Any, Callable, Dict, List, Optional

import customtkinter as ctk

# Handle branding imports with fallback
try:
    from ..branding import COLORS, FONTS
except ImportError:
    parent_dir = str(Path(__file__).parent.parent)
    if parent_dir not in sys.path:
        sys.path.insert(0, parent_dir)
    from branding import COLORS, FONTS


class TreeviewResultsPanel(ctk.CTkFrame):
    """Professional results table using Tkinter Treeview with sorting and filtering."""

    def __init__(self, parent, on_status_callback: Optional[Callable] = None):
        super().__init__(parent, fg_color=COLORS["background"])
        self.on_status = on_status_callback or (lambda x: None)
        self.results: List[Dict[str, Any]] = []
        self.filtered_results: List[Dict[str, Any]] = []
        self.sort_column: str = "filename"
        self.sort_ascending: bool = True
        self.quick_filter_text: str = ""
        self.is_cached: bool = False

        self._build_ui()

    def _build_ui(self):
        """Build results panel with Treeview."""
        # Header frame with title and stats
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(fill="x", padx=15, pady=(15, 10))

        self.title_label = ctk.CTkLabel(
            header_frame,
            text="Results",
            font=FONTS["heading"],
            text_color=COLORS["primary"]
        )
        self.title_label.pack(side="left")

        self.count_label = ctk.CTkLabel(
            header_frame,
            text="",
            font=FONTS["body"],
            text_color=COLORS["text_secondary"]
        )
        self.count_label.pack(side="left", padx=(20, 0))

        self.cache_label = ctk.CTkLabel(
            header_frame,
            text="",
            font=("Arial", 9),
            text_color=COLORS["text_secondary"]
        )
        self.cache_label.pack(side="right", padx=(10, 0))

        # Export button
        self.export_button = ctk.CTkButton(
            header_frame,
            text="📥 Export",
            font=("Arial", 9),
            height=24,
            width=80,
            fg_color=COLORS["accent"],
            text_color=COLORS["background"],
            hover_color="#E01670",
            command=self._show_export_menu,
            state="disabled"
        )
        self.export_button.pack(side="right", padx=(0, 10))

        # Quick filter frame
        filter_frame = ctk.CTkFrame(self, fg_color="transparent")
        filter_frame.pack(fill="x", padx=15, pady=(0, 10))

        ctk.CTkLabel(
            filter_frame,
            text="Quick Filter:",
            font=("Arial", 9),
            text_color=COLORS["text_primary"]
        ).pack(side="left", padx=(0, 5))

        self.quick_filter_entry = ctk.CTkEntry(
            filter_frame,
            placeholder_text="Search results by filename or path...",
            height=28,
            font=("Arial", 9),
            fg_color=COLORS["surface"],
            text_color=COLORS["text_secondary"],
            border_color=COLORS["border"],
            border_width=1
        )
        self.quick_filter_entry.pack(side="left", fill="x", expand=True)
        self.quick_filter_entry.bind("<KeyRelease>", self._on_quick_filter_change)

        # Clear filter button
        ctk.CTkButton(
            filter_frame,
            text="✕ Clear",
            width=60,
            height=28,
            font=("Arial", 9),
            fg_color=COLORS["accent"],
            text_color=COLORS["background"],
            hover_color="#E01670",
            command=self._clear_quick_filter
        ).pack(side="left", padx=(5, 0))

        # Create Treeview table
        table_frame = ctk.CTkFrame(self, fg_color=COLORS["surface"], border_width=1, border_color=COLORS["border"])
        table_frame.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        # Configure Treeview style
        self._configure_treeview_style()

        # Create Treeview
        self.tree = ttk.Treeview(
            table_frame,
            columns=("filename", "path", "modified", "type"),
            height=20,
            show="headings",
            style="Treeview"
        )

        # Define columns
        self.tree.column("filename", width=200, anchor="w", minwidth=100)
        self.tree.column("path", width=400, anchor="w", minwidth=150)
        self.tree.column("modified", width=120, anchor="center", minwidth=100)
        self.tree.column("type", width=80, anchor="center", minwidth=60)

        # Define headings with click-to-sort
        self.tree.heading("filename", text="Filename", command=lambda: self._sort_by_column("filename"))
        self.tree.heading("path", text="Path", command=lambda: self._sort_by_column("path"))
        self.tree.heading("modified", text="Modified", command=lambda: self._sort_by_column("modified"))
        self.tree.heading("type", text="Type", command=lambda: self._sort_by_column("type"))

        # Add scrollbars
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        # Grid layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Bind right-click for context menu
        self.tree.bind("<Button-3>", self._show_context_menu)
        self.tree.bind("<Button-1>", self._on_row_select)

    def _configure_treeview_style(self):
        """Configure Treeview colors with strong contrast."""
        style = ttk.Style()

        # Configure Treeview rows
        style.configure(
            "Treeview",
            background=COLORS["surface"],
            foreground=COLORS["text_primary"],
            fieldbackground=COLORS["surface"],
            borderwidth=0,
            font=("Arial", 10)
        )

        # Configure heading with strong contrast
        style.configure(
            "Treeview.Heading",
            background="#E8E8E8",   # Light gray background
            foreground="#000000",    # Black text for readability
            borderwidth=2,
            font=("Arial", 10, "bold"),
            relief="raised"
        )

        # Map hover effect on heading
        style.map(
            "Treeview.Heading",
            background=[("active", "#D0D0D0")],  # Darker gray on hover
            foreground=[("active", "#000000")]   # Black text on hover
        )

        # Map colors for row selection
        style.map(
            "Treeview",
            background=[("selected", "#0052CC")],  # Medium blue for selection
            foreground=[("selected", "#FFFFFF")]   # White text when selected
        )

        # Configure row striping
        style.configure("oddrow", background=COLORS["surface"], font=("Arial", 10))
        style.configure("evenrow", background="#F0F0F5", font=("Arial", 10))

    def _sort_by_column(self, column: str):
        """Sort results by clicking column header."""
        # Toggle direction if same column
        if self.sort_column == column:
            self.sort_ascending = not self.sort_ascending
        else:
            self.sort_column = column
            self.sort_ascending = True

        # Re-display with new sort
        if self.results:
            self._refresh_display()

    def _on_quick_filter_change(self, event=None):
        """Filter results as user types."""
        self.quick_filter_text = self.quick_filter_entry.get().strip().lower()

        if self.quick_filter_text:
            self.filtered_results = [
                r for r in self.results
                if self.quick_filter_text in r.get("filename", "").lower()
                   or self.quick_filter_text in r.get("path", "").lower()
            ]
        else:
            self.filtered_results = self.results

        if self.results:
            self._refresh_display()

    def _clear_quick_filter(self):
        """Clear filter and show all results."""
        self.quick_filter_entry.delete(0, "end")
        self.quick_filter_text = ""
        self.filtered_results = self.results

        if self.results:
            self._refresh_display()

        self.on_status("Filter cleared")

    def _refresh_display(self):
        """Refresh Treeview with current results."""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Sort results
        sorted_results = sorted(
            self.filtered_results,
            key=lambda x: x.get(self.sort_column, "").lower() if isinstance(x.get(self.sort_column), str) else x.get(self.sort_column, 0),
            reverse=not self.sort_ascending
        )

        # Add sorted results to tree
        for i, file_info in enumerate(sorted_results):
            values = (
                file_info.get("filename", ""),
                file_info.get("path", ""),
                file_info.get("modified", "")[:10],  # Just date part
                file_info.get("type", "")
            )

            # Alternate row colors for readability
            tag = "oddrow" if i % 2 == 0 else "evenrow"
            self.tree.insert("", "end", values=values, tags=(tag,))

        # Update counts
        if self.quick_filter_text:
            self.count_label.configure(text=f"{len(self.filtered_results)}/{len(self.results)} files")
        else:
            self.count_label.configure(text=f"{len(self.results)} files")

        # Update cache indicator
        if self.is_cached:
            self.cache_label.configure(text="📦 Cached results")
        else:
            self.cache_label.configure(text="🔍 Live search")

    def _on_row_select(self, event):
        """Handle row selection."""
        selection = self.tree.selection()
        if selection:
            item = selection[0]
            values = self.tree.item(item)["values"]
            self.on_status(f"Selected: {values[0]}")

    def _show_context_menu(self, event):
        """Show right-click context menu."""
        item = self.tree.selection()
        if not item:
            return

        item = item[0]
        values = self.tree.item(item)["values"]

        if len(values) < 2:
            return

        filename = values[0]
        filepath = values[1]

        if not os.path.exists(filepath):
            self.on_status("File no longer exists")
            return

        # Create context menu
        context_menu = tk.Menu(
            self.tree,
            tearoff=False,
            bg=COLORS["surface"],
            fg=COLORS["text_primary"],
            activebackground=COLORS["primary"],
            activeforeground=COLORS["background"]
        )

        context_menu.add_command(
            label="📋 Copy Full Path",
            command=lambda: self._copy_to_clipboard(filepath)
        )
        context_menu.add_command(
            label="📄 Copy Filename",
            command=lambda: self._copy_to_clipboard(filename)
        )
        context_menu.add_separator()
        context_menu.add_command(
            label="📂 Open File Location",
            command=lambda: self._open_folder(filepath)
        )
        context_menu.add_command(
            label="▶️ Open File",
            command=lambda: self._open_file(filepath)
        )
        context_menu.add_separator()
        context_menu.add_command(
            label="ℹ️ File Properties",
            command=lambda: self._show_file_info(filepath)
        )

        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def _copy_to_clipboard(self, text: str):
        """Copy text to clipboard."""
        try:
            self.tree.clipboard_clear()
            self.tree.clipboard_append(text)
            self.tree.update()
            self.on_status(f"Copied: {text[:50]}...")
        except Exception as e:
            self.on_status(f"Copy error: {e}")

    def _open_file(self, filepath: str):
        """Open file with default application."""
        try:
            if os.name == "nt":  # Windows
                os.startfile(filepath)
            elif os.name == "posix":  # macOS/Linux
                subprocess.Popen(["open" if sys.platform == "darwin" else "xdg-open", filepath])
            self.on_status(f"Opened: {filepath}")
        except Exception as e:
            self.on_status(f"Error opening file: {e}")

    def _open_folder(self, filepath: str):
        """Open file's containing folder."""
        try:
            folder = os.path.dirname(filepath)
            if os.name == "nt":
                os.startfile(folder)
            elif os.name == "posix":
                subprocess.Popen(["open" if sys.platform == "darwin" else "xdg-open", folder])
            self.on_status(f"Opened folder: {folder}")
        except Exception as e:
            self.on_status(f"Error opening folder: {e}")

    def _show_file_info(self, filepath: str):
        """Show file properties in status."""
        try:
            stat = os.stat(filepath)
            size_mb = stat.st_size / (1024 * 1024)
            from datetime import datetime
            mod_time = datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")

            info = f"File: {Path(filepath).name} | Size: {size_mb:.2f} MB | Modified: {mod_time}"
            self.on_status(info)
        except Exception as e:
            self.on_status(f"Error reading file info: {e}")

    def display_results(self, results: List[Dict[str, Any]], is_cached: bool = False):
        """Display search results in Treeview."""
        self.results = results
        self.filtered_results = results
        self.is_cached = is_cached

        # Clear filter
        self.quick_filter_entry.delete(0, "end")
        self.quick_filter_text = ""

        # Enable export button if results exist
        if results:
            self.export_button.configure(state="normal")
        else:
            self.export_button.configure(state="disabled")

        # Refresh display
        self._refresh_display()

    def display_analysis_result(self, result: str):
        """Display AI analysis result as text."""
        # Clear tree
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Add analysis text to a single cell
        self.tree.insert("", "end", values=(result, "", "", ""))
        self.title_label.configure(text="Analysis Results")
        self.count_label.configure(text=f"{result.count(chr(10))} lines")

    def _show_export_menu(self):
        """Show export format menu."""
        if not self.results:
            self.on_status("No results to export")
            return

        export_menu = tk.Menu(
            self.export_button,
            tearoff=False,
            bg=COLORS["surface"],
            fg=COLORS["text_primary"],
            activebackground=COLORS["primary"],
            activeforeground=COLORS["background"]
        )

        export_menu.add_command(
            label="📊 Export to CSV",
            command=self._export_to_csv
        )
        export_menu.add_command(
            label="📈 Export to Excel",
            command=self._export_to_excel
        )

        try:
            export_menu.tk_popup(
                self.export_button.winfo_rootx(),
                self.export_button.winfo_rooty() + self.export_button.winfo_height()
            )
        finally:
            export_menu.grab_release()

    def _export_to_csv(self):
        """Export results to CSV file."""
        if not self.results:
            self.on_status("No results to export")
            return

        # Generate default filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        default_filename = f"search_results_{timestamp}.csv"

        # Show save dialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            initialfile=default_filename,
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )

        if not file_path:
            return

        try:
            # Prepare data: use filtered results if filter is active, else all results
            export_data = self.filtered_results if self.quick_filter_text else self.results

            # Write CSV
            with open(file_path, "w", newline="", encoding="utf-8") as csvfile:
                fieldnames = ["Filename", "Path", "Modified", "Type"]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for result in export_data:
                    writer.writerow({
                        "Filename": result.get("filename", ""),
                        "Path": result.get("path", ""),
                        "Modified": result.get("modified", ""),
                        "Type": result.get("type", "")
                    })

            self.on_status(f"✓ Exported {len(export_data)} results to CSV: {Path(file_path).name}")

        except Exception as e:
            self.on_status(f"Export error: {e}")

    def _export_to_excel(self):
        """Export results to Excel file."""
        if not self.results:
            self.on_status("No results to export")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            self.on_status("Excel export requires openpyxl. Using CSV instead.")
            self._export_to_csv()
            return

        # Generate default filename with timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        default_filename = f"search_results_{timestamp}.xlsx"

        # Show save dialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
        )

        if not file_path:
            return

        try:
            # Prepare data: use filtered results if filter is active, else all results
            export_data = self.filtered_results if self.quick_filter_text else self.results

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Search Results"

            # Write header with formatting
            headers = ["Filename", "Path", "Modified", "Type"]
            ws.append(headers)

            # Format header row
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Write data
            for result in export_data:
                ws.append([
                    result.get("filename", ""),
                    result.get("path", ""),
                    result.get("modified", ""),
                    result.get("type", "")
                ])

            # Adjust column widths
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 12

            # Save workbook
            wb.save(file_path)
            self.on_status(f"✓ Exported {len(export_data)} results to Excel: {Path(file_path).name}")

        except Exception as e:
            self.on_status(f"Export error: {e}")

    def clear_results(self):
        """Clear all results."""
        for item in self.tree.get_children():
            self.tree.delete(item)

        self.count_label.configure(text="")
        self.cache_label.configure(text="")
        self.title_label.configure(text="Results")
        self.results = []
        self.filtered_results = []
        self.export_button.configure(state="disabled")
